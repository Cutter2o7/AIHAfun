"""Utility functions for fetching Daily Dose videos and translations."""

from __future__ import annotations

import os
import re
import requests
import shutil
import subprocess
import webbrowser
import warnings
from pathlib import Path

try:
    import uno  # LibreOffice UNO API for Calc automation
except ImportError:  # noqa: W0707
    uno = None

from dotenv import load_dotenv
from googleapiclient.discovery import build

load_dotenv()

# Mapping of Bible books to numerical codes used by the IQ Bible API
BOOK_CODES = {
    "Genesis": 1,
    "Exodus": 2,
    "Leviticus": 3,
    "Numbers": 4,
    "Deuteronomy": 5,
    "Joshua": 6,
    "Judges": 7,
    "Ruth": 8,
    "1 Samuel": 9,
    "2 Samuel": 10,
    "1 Kings": 11,
    "2 Kings": 12,
    "1 Chronicles": 13,
    "2 Chronicles": 14,
    "Ezra": 15,
    "Nehemiah": 16,
    "Esther": 17,
    "Job": 18,
    "Psalm": 19,
    "Proverbs": 20,
    "Ecclesiastes": 21,
    "Song of Solomon": 22,
    "Song of Songs": 22,
    "Isaiah": 23,
    "Jeremiah": 24,
    "Lamentations": 25,
    "Ezekiel": 26,
    "Daniel": 27,
    "Hosea": 28,
    "Joel": 29,
    "Amos": 30,
    "Obadiah": 31,
    "Jonah": 32,
    "Micah": 33,
    "Nahum": 34,
    "Habakkuk": 35,
    "Zephaniah": 36,
    "Haggai": 37,
    "Zechariah": 38,
    "Malachi": 39,
    "Matthew": 40,
    "Mark": 41,
    "Luke": 42,
    "John": 43,
    "Acts": 44,
    "Romans": 45,
    "1 Corinthians": 46,
    "2 Corinthians": 47,
    "Galatians": 48,
    "Ephesians": 49,
    "Philippians": 50,
    "Colossians": 51,
    "1 Thessalonians": 52,
    "2 Thessalonians": 53,
    "1 Timothy": 54,
    "2 Timothy": 55,
    "Titus": 56,
    "Philemon": 57,
    "Hebrews": 58,
    "James": 59,
    "1 Peter": 60,
    "2 Peter": 61,
    "1 John": 62,
    "2 John": 63,
    "3 John": 64,
    "Jude": 65,
    "Revelation": 66,
}

# Reverse mapping from numerical book codes back to names
BOOK_NAMES = {value: key for key, value in BOOK_CODES.items()}


def slug_from_verse_id(verse_id: str | int) -> str | None:
    """Return a ``reference_slug`` style string from a numeric verse ID."""
    s = str(verse_id)
    if len(s) < 7:
        return None
    try:
        book_id = int(s[:-6])
        chapter = int(s[-6:-3])
        verse = int(s[-3:])
    except ValueError:
        return None

    book = BOOK_NAMES.get(book_id)
    if not book:
        return None
    return f"{book} {chapter}_{verse}"


def parse_reference(title: str) -> str | None:
    """Extract verse information from a video title."""
    pattern = r"([1-3]?\s?[A-Za-z ]+?)\s+(\d+):(\d+)"
    match = re.search(pattern, title)
    if not match:
        return None

    book_name = match.group(1).strip()
    # Normalize spacing and capitalization for lookup
    book_key = " ".join(word.capitalize() for word in book_name.split())
    book_id = BOOK_CODES.get(book_key)
    if book_id is None:
        return None
    chapter = int(match.group(2))
    verse = int(match.group(3))
    return f"{book_id}{chapter:03d}{verse:03d}"


def reference_slug(title: str) -> str | None:
    """Return the verse reference from ``title`` formatted for filenames."""
    pattern = r"([1-3]?\s?[A-Za-z ]+?)\s+(\d+):(\d+)"
    match = re.search(pattern, title)
    if not match:
        return None
    book = match.group(1).strip()
    chapter = match.group(2)
    verse = match.group(3)
    return f"{book} {chapter}_{verse}"


# --- Spreadsheet helpers ---------------------------------------------------

def open_translation_spreadsheet(env_var: str, ref_slug: str | None = None) -> Path | None:
    """Copy the spreadsheet referenced by ``env_var`` and open it."""
    dest = prepare_translation_spreadsheet(env_var, ref_slug)
    if dest is None:
        return None
    open_spreadsheet(dest)
    return dest


def prepare_translation_spreadsheet(env_var: str, ref_slug: str | None = None) -> Path | None:
    """Copy the spreadsheet referenced by ``env_var`` to ``~/Desktop/Daily Dose``."""
    path = os.getenv(env_var)
    if not path:
        print(f"{env_var} not configured in .env")
        return None
    src = Path(path).expanduser()
    if not src.is_file():
        print(f"Translation file not found: {src}")
        return None
    dest_dir = Path.home() / "Desktop" / "Daily Dose"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / (f"{ref_slug}{src.suffix}" if ref_slug else src.name)
    try:
        shutil.copy(src, dest)
    except Exception as err:  # noqa: BLE001
        print(f"Failed to copy translation file: {err}")
        return None
    return dest


def open_spreadsheet(path: Path) -> None:
    """Open an existing spreadsheet with LibreOffice."""
    if uno is not None:
        try:
            local_ctx = uno.getComponentContext()
            resolver = local_ctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.bridge.UnoUrlResolver", local_ctx
            )
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext"
            )
            desktop_srv = ctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.frame.Desktop", ctx
            )
            desktop_srv.loadComponentFromURL(Path(path).as_uri(), "_blank", 0, tuple())
            return
        except Exception as err:  # noqa: BLE001
            print(f"UNO open failed: {err}")

    try:
        libreoffice_exe = os.getenv("LIBREOFFICE_EXE", "libreoffice")
        subprocess.Popen([libreoffice_exe, "--calc", str(path)])
    except Exception as err:  # noqa: BLE001
        print(f"Failed to open translation file {path}: {err}")


def write_words_to_spreadsheet(file_path: Path, words: list[str], start_row: int = 4) -> None:
    """Write the given list of words to column A of a spreadsheet."""
    try:
        from openpyxl import load_workbook
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
        )
    except Exception as err:  # noqa: BLE001
        print(f"openpyxl not available: {err}")
        return

    try:
        wb = load_workbook(file_path)
    except Exception as err:  # noqa: BLE001
        print(f"Unable to open spreadsheet {file_path}: {err}")
        return

    ws = wb.active
    row = start_row
    for word in words:
        ws.cell(row=row, column=1, value=word)
        row += 1

    try:
        wb.save(file_path)
    except Exception as err:  # noqa: BLE001
        print(f"Unable to save spreadsheet {file_path}: {err}")


# --- Main Daily Dose retrieval ---------------------------------------------

def fetch_daily_dose(language: str) -> None:
    """Fetch and display the latest Daily Dose video for ``language``."""
    lang = language.lower()
    if lang == "hebrew":
        query = "Daily Dose of Hebrew"
        env_var = "HEBREW_TRANSLATION_FILE"
    elif lang == "greek":
        query = "Daily Dose of Greek"
        env_var = "GREEK_TRANSLATION_FILE"
    else:
        raise ValueError("language must be 'hebrew' or 'greek'")

    api_key = os.getenv("YOUTUBE_TOKEN")
    if not api_key:
        print("YOUTUBE_TOKEN is not set in the environment")
        return

    try:
        youtube = build("youtube", "v3", developerKey=api_key)

        channel_search = (
            youtube.search()
            .list(q=query, type="channel", part="id", maxResults=1)
            .execute()
        )
        channel_id = channel_search["items"][0]["id"]["channelId"]

        video_search = (
            youtube.search()
            .list(
                channelId=channel_id,
                part="id,snippet",
                order="date",
                maxResults=1,
                type="video",
            )
            .execute()
        )

        item = video_search["items"][0]
        video_id = item["id"]["videoId"]
        title = item["snippet"]["title"]
        url = f"https://www.youtube.com/watch?v={video_id}"
        print(f"{query}: {title}\n{url}")

        verse_id = parse_reference(title)
        ref_slug = reference_slug(title)
        if ref_slug is None and verse_id:
            ref_slug = slug_from_verse_id(verse_id)

        spreadsheet_path = prepare_translation_spreadsheet(env_var, ref_slug)
        if verse_id:
            api_key_rapid = os.getenv("RAPIDAPI_KEY")
            api_host = os.getenv("RAPIDAPI_HOST")
            if api_key_rapid and api_host:
                headers = {
                    "x-rapidapi-key": api_key_rapid,
                    "x-rapidapi-host": api_host,
                }
                try:
                    resp = requests.get(
                        "https://iq-bible.p.rapidapi.com/GetOriginalText",
                        headers=headers,
                        params={"verseId": str(verse_id)},
                        timeout=10,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    words_sorted = sorted(data, key=lambda x: int(x.get("orig_order", 0)))
                    words = [entry.get("word", "") for entry in words_sorted]
                    if spreadsheet_path:
                        write_words_to_spreadsheet(spreadsheet_path, words)
                        open_spreadsheet(spreadsheet_path)
                except Exception as err:  # noqa: BLE001
                    print(f"Failed to fetch verse text: {err}")
            else:
                print("RAPIDAPI credentials not set in environment")
        else:
            print("Unable to parse verse reference from title")

        webbrowser.open(url)
    except Exception as err:  # noqa: BLE001
        print(f"Failed to retrieve {query} video: {err}")
