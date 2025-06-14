# === 1. Imports and Configuration ===
import os
import re
import requests
import time
import shutil
import subprocess
import webbrowser
import warnings
from pathlib import Path

try:
    import uno  # LibreOffice UNO API for Calc automation
except ImportError:  # noqa: W0707
    uno = None
from dotenv import load_dotenv
from googleapiclient.discovery import build
load_dotenv()
HEBREW_TRANSLATION_FILE = os.getenv("HEBREW_TRANSLATION_FILE")
GREEK_TRANSLATION_FILE = os.getenv("GREEK_TRANSLATION_FILE")


# Mapping of Bible books to numerical codes used by the IQ Bible API
BOOK_CODES = {
    "Genesis": 1,
    "Exodus": 2,
    "Leviticus": 3,
    "Numbers": 4,
    "Deuteronomy": 5,
    "Joshua": 6,
    "Judges": 7,
    "Ruth": 8,
    "1 Samuel": 9,
    "2 Samuel": 10,
    "1 Kings": 11,
    "2 Kings": 12,
    "1 Chronicles": 13,
    "2 Chronicles": 14,
    "Ezra": 15,
    "Nehemiah": 16,
    "Esther": 17,
    "Job": 18,
    "Psalms": 19,
    "Proverbs": 20,
    "Ecclesiastes": 21,
    "Song of Solomon": 22,
    "Song of Songs": 22,
    "Isaiah": 23,
    "Jeremiah": 24,
    "Lamentations": 25,
    "Ezekiel": 26,
    "Daniel": 27,
    "Hosea": 28,
    "Joel": 29,
    "Amos": 30,
    "Obadiah": 31,
    "Jonah": 32,
    "Micah": 33,
    "Nahum": 34,
    "Habakkuk": 35,
    "Zephaniah": 36,
    "Haggai": 37,
    "Zechariah": 38,
    "Malachi": 39,
    "Matthew": 40,
    "Mark": 41,
    "Luke": 42,
    "John": 43,
    "Acts": 44,
    "Romans": 45,
    "1 Corinthians": 46,
    "2 Corinthians": 47,
    "Galatians": 48,
    "Ephesians": 49,
    "Philippians": 50,
    "Colossians": 51,
    "1 Thessalonians": 52,
    "2 Thessalonians": 53,
    "1 Timothy": 54,
    "2 Timothy": 55,
    "Titus": 56,
    "Philemon": 57,
    "Hebrews": 58,
    "James": 59,
    "1 Peter": 60,
    "2 Peter": 61,
    "1 John": 62,
    "2 John": 63,
    "3 John": 64,
    "Jude": 65,
    "Revelation": 66,
}

def parse_reference(title):
    """Extract verse information from a video title."""
    pattern = r"([1-3]?\s?[A-Za-z ]+?)\s+(\d+):(\d+)"
    match = re.search(pattern, title)
    if not match:
        return None

    book_name = match.group(1).strip()
    # Normalize spacing and capitalization for lookup
    book_key = " ".join(word.capitalize() for word in book_name.split())
    book_id = BOOK_CODES.get(book_key)
    if book_id is None:
        return None
    chapter = int(match.group(2))
    verse = int(match.group(3))
    return f"{book_id}{chapter:03d}{verse:03d}"
    



def open_translation_spreadsheet(env_var):
    """Copy the spreadsheet referenced by ``env_var`` to the Desktop and open
    it with LibreOffice.

    This helper does not require Microsoft Excel. The workbook should be in
    ``.xlsx`` format so that ``openpyxl`` can update it and LibreOffice Calc can
    open it. The function returns the path to the copied file on success,
    otherwise ``None``.
    """
    dest = prepare_translation_spreadsheet(env_var)
    if dest is None:
        return None
    open_spreadsheet(dest)
    return dest


def prepare_translation_spreadsheet(env_var):
    """Copy the spreadsheet referenced by ``env_var`` to the Desktop."""
    path = os.getenv(env_var)
    if not path:
        print(f"{env_var} not configured in .env")
        return None
    src = Path(path).expanduser()
    if not src.is_file():
        print(f"Translation file not found: {src}")
        return None
    desktop = Path.home() / "Desktop"
    desktop.mkdir(parents=True, exist_ok=True)
    dest = desktop / src.name
    try:
        shutil.copy(src, dest)
    except Exception as err:
        print(f"Failed to copy translation file: {err}")
        return None
    return dest


def open_spreadsheet(path):
    """Open an existing spreadsheet with LibreOffice."""
    if uno is not None:
        try:
            local_ctx = uno.getComponentContext()
            resolver = local_ctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.bridge.UnoUrlResolver", local_ctx
            )
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext"
            )
            desktop_srv = ctx.ServiceManager.createInstanceWithContext(
                "com.sun.star.frame.Desktop", ctx
            )
            desktop_srv.loadComponentFromURL(Path(path).as_uri(), "_blank", 0, tuple())
            return
        except Exception as err:
            print(f"UNO open failed: {err}")

    try:
        libreoffice_exe = os.getenv("LIBREOFFICE_EXE", "libreoffice")
        subprocess.Popen([libreoffice_exe, "--calc", str(path)])
    except Exception as err:
        print(f"Failed to open translation file {path}: {err}")


def write_words_to_spreadsheet(file_path, words, start_row=4):
    """Write the given list of words to column A of a spreadsheet.

    The file must be in ``.xlsx`` format so that ``openpyxl`` can modify it.
    LibreOffice users can open and edit the resulting workbook without
    Microsoft Excel.
    """
    try:
        from openpyxl import load_workbook
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
        )
    except Exception as err:
        print(f"openpyxl not available: {err}")
        return

    try:
        wb = load_workbook(file_path)
    except Exception as err:
        print(f"Unable to open spreadsheet {file_path}: {err}")
        return

    ws = wb.active
    row = start_row
    for word in words:
        ws.cell(row=row, column=1, value=word)
        row += 1

    try:
        wb.save(file_path)
    except Exception as err:
        print(f"Unable to save spreadsheet {file_path}: {err}")
# === 2. Fetch Content ===
def fetch_daily_dose_hebrew():
    """Fetch and display the latest Daily Dose of Hebrew video."""
    # Load API token from .env
    load_dotenv()
    api_key = os.getenv("YOUTUBE_TOKEN")
    if not api_key:
        print("YOUTUBE_TOKEN is not set in the environment")
        return

    # Initialize YouTube client
    try:
        youtube = build("youtube", "v3", developerKey=api_key)

        # Determine channel ID for Daily Dose of Hebrew
        channel_search = (
            youtube.search()
            .list(q="Daily Dose of Hebrew", type="channel", part="id", maxResults=1)
            .execute()
        )
        channel_id = channel_search["items"][0]["id"]["channelId"]

        # Get the most recent video from the channel
        video_search = (
            youtube.search()
            .list(
                channelId=channel_id,
                part="id,snippet",
                order="date",
                maxResults=1,
                type="video",
            )
            .execute()
        )

        item = video_search["items"][0]
        video_id = item["id"]["videoId"]
        title = item["snippet"]["title"]
        url = f"https://www.youtube.com/watch?v={video_id}"

        print(f"Daily Dose of Hebrew: {title}\n{url}")

        # Copy today's translation spreadsheet but don't open it yet
        spreadsheet_path = prepare_translation_spreadsheet("HEBREW_TRANSLATION_FILE")


        verse_id = parse_reference(title)
        if verse_id:
            api_key_rapid = os.getenv("RAPIDAPI_KEY")
            api_host = os.getenv("RAPIDAPI_HOST")
            if api_key_rapid and api_host:
                headers = {
                    "x-rapidapi-key": api_key_rapid,
                    "x-rapidapi-host": api_host,
                }
                try:
                    resp = requests.get(
                        "https://iq-bible.p.rapidapi.com/GetOriginalText",
                        headers=headers,
                        params={"verseId": str(verse_id)},
                        timeout=10,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    words_sorted = sorted(
                        data,
                        key=lambda x: int(x.get("orig_order", 0))
                    )
                    words = [entry.get("word", "") for entry in words_sorted]
                    if spreadsheet_path:
                        write_words_to_spreadsheet(spreadsheet_path, words)
                        open_spreadsheet(spreadsheet_path)
                except Exception as err:
                    print(f"Failed to fetch verse text: {err}")
            else:
                print("RAPIDAPI credentials not set in environment")
        else:
            print("Unable to parse verse reference from title")

        webbrowser.open(url)
    except Exception as err:
        print(f"Failed to retrieve Daily Dose of Hebrew video: {err}")

def fetch_daily_dose_greek():
    """Fetch and display the latest Daily Dose of Greek video."""
    # Load API token from .env
    load_dotenv()
    api_key = os.getenv("YOUTUBE_TOKEN")
    if not api_key:
        print("YOUTUBE_TOKEN is not set in the environment")
        return

    # Initialize YouTube client
    try:
        youtube = build("youtube", "v3", developerKey=api_key)

        # Determine channel ID for Daily Dose of Greek
        channel_search = (
            youtube.search()
            .list(q="Daily Dose of Greek", type="channel", part="id", maxResults=1)
            .execute()
        )
        channel_id = channel_search["items"][0]["id"]["channelId"]

        # Get the most recent video from the channel
        video_search = (
            youtube.search()
            .list(
                channelId=channel_id,
                part="id,snippet",
                order="date",
                maxResults=1,
                type="video",
            )
            .execute()
        )

        item = video_search["items"][0]
        video_id = item["id"]["videoId"]
        title = item["snippet"]["title"]
        url = f"https://www.youtube.com/watch?v={video_id}"

        print(f"Daily Dose of Greek: {title}\n{url}")

        # Copy today's translation spreadsheet but don't open it yet
        spreadsheet_path = prepare_translation_spreadsheet("GREEK_TRANSLATION_FILE")

        verse_id = parse_reference(title)
        if verse_id:
            api_key_rapid = os.getenv("RAPIDAPI_KEY")
            api_host = os.getenv("RAPIDAPI_HOST")
            if api_key_rapid and api_host:
                headers = {
                    "x-rapidapi-key": api_key_rapid,
                    "x-rapidapi-host": api_host,
                }
                try:
                    resp = requests.get(
                        "https://iq-bible.p.rapidapi.com/GetOriginalText",
                        headers=headers,
                        params={"verseId": str(verse_id)},
                        timeout=10,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    words_sorted = sorted(
                        data,
                        key=lambda x: int(x.get("orig_order", 0))
                    )
                    words = [entry.get("word", "") for entry in words_sorted]
                    if spreadsheet_path:
                        write_words_to_spreadsheet(spreadsheet_path, words)
                        open_spreadsheet(spreadsheet_path)
                except Exception as err:
                    print(f"Failed to fetch verse text: {err}")
            else:
                print("RAPIDAPI credentials not set in environment")
        else:
            print("Unable to parse verse reference from title")

        webbrowser.open(url)
    except Exception as err:
        print(f"Failed to retrieve Daily Dose of Greek video: {err}")

def open_bible_study_tools():
    """Open Bible tools, translation websites, and LibreOffice study doc."""
    pass

# === 3. Timers for Language and Study ===
def start_timer(name, minutes):
    """Start a named countdown timer with a Tkinter progress bar."""
    import tkinter as tk
    from tkinter import ttk

    total_seconds = int(minutes * 60)

    root = tk.Tk()
    root.title(f"{name} Timer")

    progress_var = tk.DoubleVar(value=0)
    progress = ttk.Progressbar(
        root,
        maximum=total_seconds,
        length=300,
        variable=progress_var,
    )
    progress.pack(padx=20, pady=10)

    label = tk.Label(root, text=f"Time remaining: {minutes:02d}:00")
    label.pack()

    def update():
        nonlocal total_seconds
        mins, secs = divmod(total_seconds, 60)
        label.config(text=f"Time remaining: {mins:02d}:{secs:02d}")
        progress_var.set((minutes * 60) - total_seconds)
        if total_seconds > 0:
            total_seconds -= 1
            root.after(1000, update)
        else:
            label.config(text=f"{name} complete!")
            progress_var.set(minutes * 60)
            root.after(2000, root.destroy)

    def skip_timer():
        nonlocal total_seconds
        total_seconds = 0
        update()

    skip_button = tk.Button(root, text="Skip", command=skip_timer)
    skip_button.pack(pady=5)

    update()
    root.mainloop()

def run_study_timers():
    """Run timers for Hebrew, Greek, and Bible study sessions."""
    start_timer("Hebrew", 10)
    start_timer("Greek", 10)
    start_timer("Bible Study", 30)

# === 4. Weather ===
def fetch_weather():
    """Fetch and display the hourly weather forecast for the next two days."""
    from datetime import datetime, timezone, timedelta

    # Coordinates for 32°41'00.8"N 97°24'51.2"W
    latitude = 32.683556
    longitude = -97.414222

    # Step 1: Get the hourly forecast URL from the /points endpoint
    points_url = f"https://api.weather.gov/points/{latitude},{longitude}"
    try:
        response = requests.get(points_url, timeout=10)
        response.raise_for_status()
        forecast_url = response.json()["properties"]["forecastHourly"]
    except (requests.RequestException, KeyError) as err:
        print(f"Unable to determine hourly forecast URL: {err}")
        return

    # Step 2: Retrieve the hourly forecast
    try:
        forecast_resp = requests.get(forecast_url, timeout=10)
        forecast_resp.raise_for_status()
        periods = forecast_resp.json()["properties"]["periods"]
    except (requests.RequestException, KeyError) as err:
        print(f"Unable to fetch forecast data: {err}")
        return

    # Filter to the next two days
    now = datetime.now(timezone.utc)
    cutoff = now + timedelta(days=2)

    print("\nHourly Forecast (next 48 hours):")
    for period in periods:
        start_time_str = period.get("startTime")
        if not start_time_str:
            continue
        try:
            start_time = datetime.fromisoformat(start_time_str)
        except ValueError:
            # If timezone info is missing, assume UTC
            start_time = datetime.fromisoformat(start_time_str + "+00:00")
        if start_time >= cutoff:
            break

        name = period.get("name") or start_time.strftime("%a %I %p")
        short = period.get("shortForecast")
        temp = period.get("temperature")
        unit = period.get("temperatureUnit", "")
        print(f"{name}: {temp}{unit}, {short}")

# === 5. Novel Writing Prompt ===
def prompt_novel_scene_writing():
    """Display prompt and open tools for novel writing."""
    print("📝 Write 2 scenes for your novel today!")
    open_novel_resources()
    start_timer("Scene Writing", 15)

def open_novel_resources():
    """Open writing website and LibreOffice document for scene writing."""
    pass

# === 6. Image Generation Prompt ===
def prompt_image_generation():
    """Display prompt and open tools for AI art generation."""
    print("🎨 Generate 2 images based on your scenes or ideas!")
    open_image_resources()
    start_timer("Image Generation", 15)

def open_image_resources():
    """Open AI art website or local tool and image save folder."""
    pass

# === 7. Optional Extras ===
def display_daily_verse():
    """Fetch and display a daily Bible verse."""
    pass

def generate_todo_list():
    """Generate or display the day’s to-do list."""
    pass

# === 8. Main Routine ===
def main():

    # --- Timed study routine ---
    #start_timer("Prayer Timer", 5)

    # Daily Dose of Hebrew followed by practice
    fetch_daily_dose_hebrew()
    start_timer("Daily Hebrew Practice", 10)

    # Daily Dose of Greek followed by practice
    fetch_daily_dose_greek()
    start_timer("Daily Greek Practice", 10)

    # Open study tools then timed study session
    open_bible_study_tools()
    start_timer("Daily study", 10)
    
    # --- Post-study weather check ---
    fetch_weather()
    
    # --- Writing task ---
    #prompt_novel_scene_writing()
    
    # --- Art task ---
    #prompt_image_generation()
    
    # --- Optional extras (if enabled) ---
    #display_daily_verse()
    #generate_todo_list()

if __name__ == "__main__":
    main()
